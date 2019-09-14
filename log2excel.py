"""
将本程序的记录文本文件转存到excel工作表中。数据结构：
文件夹全名,题号,得分,批注。
错误文件记录：
错误代码，原始记录。
2019.03.12：不再兼容第一次的记录文档。
"""
source_file = r"D:\个人文件\学习\本科\第5学期\C语言助教\第02次作业\工作区2\log.txt"
excel_file = r"D:\个人文件\学习\本科\第5学期\C语言助教\第02次作业\《C程序设计》作业批改结果【第2次】.xlsx"
problem_count = 8  # 题目总数。超过这个数的题号将被忽略

# out_excel = 'source/《程序设计》-2017地海-作业批改结果 【第2次】-out.xlsx'
import openpyxl
from datetime import datetime
import re
error_log = f"output/error_log_{datetime.now().strftime('%Y-%m-%d')}.txt"

def main(source_file,excel_file,error_log,*,num_col=2,start_col=5,problem_count=8,start_time:datetime=None):
    elog = open(error_log,'a',encoding='utf-8',errors='ignore')
    elog.write(f'//{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}写入\n')
    src = open(source_file,'r',encoding='utf-8',errors='ignore')
    excel = openpyxl.load_workbook(excel_file)
    ws = excel.active

    started = False
    if start_time is None:
        started = True
    total_note_col = start_col + 2*problem_count
    for line in src:
        line = line.replace(chr(0xFEFF),'')
        if line[:2] == '//' or '//' in line: # utf-8-bom
            if not started:
                tm = datetime.strptime(line,'//打开时间：%y-%m-%d %H:%M:%S')
                if tm >= start_time:
                    started = True
            continue
        if not started:
            continue
        line = line.strip()
        if not line:
            continue
        dir_name,file_name,tm,num,marks,note = line.split(',',maxsplit=5)
        try:
            num = int(num)
        except ValueError:
            print("invalid num",line)
            num=-1
        if not num<=problem_count:
            print("非法的题号(-4):",num,line)
            elog.write(f"-4,{line}\n")
        row,col = find_pos(dir_name,file_name,int(num),ws,num_col,start_col)

        if row == -1:
            # 分支无效
            print('姓名与学号不匹配(-1):',line)
            elog.write(f"-1,{line}\n")
        elif row == -2:
            print('表中找不到学号(-2):',line)
            elog.write(f"-2,{line}\n")
        elif row == -3:
            print('不含学号的文件夹名(-3):',line)
            elog.write(f"-3,{line}\n")
        else:
            try:
                marks_int = int(marks)
            except:
                marks_int = marks
            ws.cell(row,col,value=marks_int)
            ws.cell(row,col+1,value=note)

            # 在备注一栏中标注题号对应的文件名。每条记录用||包围，方便正则查找。
            file_note = ws.cell(row, total_note_col).value
            if not file_note or not isinstance(file_note,str):
                if file_note:
                    print("strange filenote",file_note)
                file_note = "题目对应源文件|"
            newlog = make_problem_file_log(file_note,num,file_name)
            ws.cell(row,total_note_col,value=newlog)


    ws.cell(ws.max_row+1,1,
            value=f'此文件由批改程序在{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}根据记录文件写入')

    excel.save(excel_file)
    elog.close()
    src.close()

def make_problem_file_log(previous:str,pro_num:int,filename:str)->str:
    """
    生成并返回题号对应文件名记录。
    """
    pro_files = previous.split('|')[1:]
    pro_file_dict = {}
    for pro_file_log in pro_files:
        if not pro_file_log:
            continue
        num, file = pro_file_log.split(':', maxsplit=1)
        num = int(num)
        pro_file_dict[num] = file
    pro_file_dict[pro_num] = filename
    new = '题目对应源文件|'+'|'.join(map(lambda x:f"{x[0]}:{x[1]}",pro_file_dict.items()))
    return new

def find_pos(dir_name:str,file_name:str,num:int,ws,num_col,start_col)->(int,int):
    """
    根据文件夹名和题号返回【成绩格所在的行列号】。错误代码：
    若姓名学号不匹配，返回-1，-1.
    若表中找不到学号，返回-2，-2.
    若dir_name中找不到学号，返回-3，-3.
    """
    nums = re.findall(r'(\d+)',dir_name)
    nums.sort(key=len,reverse=True)
    if not nums:
        return -3,-3
    stu_num = nums[0]
    stu_row = None
    for row in range(1,ws.max_row+1):
        if str(ws.cell(row,num_col).value).strip() == stu_num:
            stu_row = row
            break
    if not stu_row:
        return -2,-2

    nameInList = str(ws.cell(stu_row,num_col+1).value).strip()
    if  nameInList not in dir_name and nameInList not in file_name:
        print('警告：姓名与学号不匹配，以学号为准录入(W-1):',nameInList,stu_num)
        # return -1,-1

    col = start_col + (num-1)*2
    return stu_row,col

def numFromDirName(name:str)->str:
    """
    从文件夹名中获取最可能的学号
    """
    nums = re.findall(r'(\d+)', name)
    nums.sort(key=len, reverse=True)
    if not nums:
        return ""
    return nums[0]

if __name__ == '__main__':
    main(source_file,excel_file,error_log,problem_count=problem_count,start_time=None)